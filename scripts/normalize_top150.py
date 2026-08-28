"""Crée une copie NORMALISÉE d'un export de ventes Cultura.

- L'original n'est jamais modifié (ouvert en lecture des valeurs calculées).
- Les EAN, souvent stockés en FORMULE Excel liée à un classeur externe, sont
  FIGÉS en valeurs texte → plus aucune dépendance à l'ancien fichier source.
- En-têtes nettoyés (ex. 'VA TTC_' -> 'VA TTC'), colonnes vides supprimées.
- Ajoute une colonne 'prix_moyen_vente_periode_ttc' (= VA TTC / Qté) — indicateur
  HISTORIQUE explicite, à ne jamais confondre avec un prix actuel.

Usage :
    python scripts/normalize_top150.py "<entrée.xlsx>" ["<sortie.xlsx>"]
"""

from __future__ import annotations

import sys
import unicodedata
from pathlib import Path

from openpyxl import Workbook, load_workbook


def _norm(h: str) -> str:
    s = unicodedata.normalize("NFKD", str(h or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.lower().replace("_", " ").split())


def _clean_header(h: str) -> str:
    return str(h or "").strip().rstrip("_").strip()


def normalize(input_path: str, output_path: str) -> None:
    src = Path(input_path)
    if not src.exists():
        raise SystemExit(f"Fichier introuvable : {src}")

    wb = load_workbook(src, data_only=True)  # valeurs calculées (EAN en cache inclus)
    sheet = next(
        (n for n in wb.sheetnames
         if any(_norm(c.value) == "ean" for c in wb[n][1])),
        wb.sheetnames[0],
    )
    ws = wb[sheet]

    raw_headers = [c.value for c in ws[1]]
    # Indices des colonnes à conserver (en-tête non vide).
    keep = [i for i, h in enumerate(raw_headers) if str(h or "").strip()]
    headers = [_clean_header(raw_headers[i]) for i in keep]

    def find(*names: str) -> int | None:
        for idx, h in enumerate(headers):
            if _norm(h) in names:
                return idx
        return None

    i_ean = find("ean")
    i_vattc = find("va ttc")
    i_qte = find("qte", "quantite")

    out = Workbook()
    ows = out.active
    ows.title = "top150"
    ows.append([*headers, "prix_moyen_vente_periode_ttc"])

    n_rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = [row[i] for i in keep]
        if all(v is None or str(v).strip() == "" for v in cells):
            continue

        # Figer l'EAN en texte (supprime toute dépendance à la formule externe).
        if i_ean is not None and cells[i_ean] is not None:
            cells[i_ean] = str(cells[i_ean]).strip().removesuffix(".0")

        # Indicateur historique explicite (jamais un prix actuel).
        pm = None
        if i_vattc is not None and i_qte is not None:
            try:
                v, q = float(cells[i_vattc]), float(cells[i_qte])
                pm = round(v / q, 2) if q else None
            except (TypeError, ValueError):
                pm = None

        ows.append([*cells, pm])
        n_rows += 1

    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out.save(out_path)
    print(f"Feuille source       : {sheet}")
    print(f"En-têtes normalisés  : {headers}")
    print(f"Lignes écrites       : {n_rows}")
    print(f"EAN figés en valeurs : {'oui' if i_ean is not None else 'colonne EAN introuvable'}")
    print(f"Copie normalisée     : {out_path}")
    print("Original inchangé     : oui (ouvert en lecture seule)")


if __name__ == "__main__":
    inp = sys.argv[1] if len(sys.argv) > 1 else None
    if not inp:
        raise SystemExit("Indique le fichier d'entrée en argument.")
    outp = sys.argv[2] if len(sys.argv) > 2 else "data/samples/top150_normalise.xlsx"
    normalize(inp, outp)

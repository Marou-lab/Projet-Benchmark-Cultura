"""Lecture d'un fichier d'entrée (Excel .xlsx ou CSV) et mapping souple des colonnes.

Le mapping normalise les en-têtes (minuscules, sans accents) puis les rapproche
d'alias FR courants. Les colonnes non reconnues sont ignorées (mais listées dans
le résultat de mapping pour transparence).
"""

from __future__ import annotations

import csv
import unicodedata
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from ..models import Condition, Product
from ..validation.ean import validate_ean

# Alias d'en-têtes -> champ canonique. Clés déjà normalisées (voir _normalize_header).
COLUMN_ALIASES: dict[str, str] = {
    "ean": "ean", "gtin": "ean", "code ean": "ean", "ean13": "ean",
    "code barre": "ean", "code barres": "ean", "code-barres": "ean", "codebarre": "ean",
    "nom": "name", "produit": "name", "designation": "name", "libelle": "name",
    "titre": "name", "intitule": "name", "nom du produit": "name",
    "marque": "brand", "brand": "brand", "fabricant": "brand",
    "categorie": "category", "category": "category", "rayon": "category",
    "univers": "category", "famille": "category",
    "prix": "price_internal", "prix cultura": "price_internal",
    "prix de vente": "price_internal", "prix ttc": "price_internal",
    "notre prix": "price_internal", "prix reference": "price_internal",
    "vendeur": "seller_internal", "vendeur cultura": "seller_internal",
    "marchand": "seller_internal", "shop": "seller_internal",
    "etat": "condition", "condition": "condition", "state": "condition",
}


@dataclass
class LoadResult:
    products: list[Product]
    mapping: dict[str, str]          # en-tête d'origine -> champ canonique
    unmapped_headers: list[str]      # en-têtes non reconnus


def _normalize_header(h: str) -> str:
    """minuscule, sans accents, espaces compactés."""
    if h is None:
        return ""
    s = unicodedata.normalize("NFKD", str(h))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.lower().replace("_", " ").split())


def parse_price(value) -> float | None:
    """Parse un prix FR : '1 299,99 €' / '1299.99' / 1299.99 -> 1299.99."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(" ", " ")
    s = s.replace("€", "").replace("EUR", "").replace("eur", "").strip()
    s = s.replace(" ", "")
    if "," in s and "." in s:          # format FR '1.299,99' -> point = millier
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _parse_condition(value) -> Condition:
    s = _normalize_header(value)
    if not s:
        return Condition.INCONNU
    if any(k in s for k in ("occasion", "reconditionne", "seconde main", "used", "refurb")):
        return Condition.OCCASION
    if "neuf" in s or "new" in s:
        return Condition.NEUF
    return Condition.INCONNU


def _read_rows(path: Path) -> list[dict[str, object]]:
    """Renvoie une liste de dicts {en-tête d'origine: valeur}."""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return []
        headers = [str(h) if h is not None else "" for h in rows[0]]
        return [dict(zip(headers, r)) for r in rows[1:]]
    if suffix in (".csv", ".tsv"):
        delim = "\t" if suffix == ".tsv" else None
        with open(path, encoding="utf-8-sig", newline="") as f:
            sample = f.read(4096)
            f.seek(0)
            if delim is None:
                try:
                    delim = csv.Sniffer().sniff(sample, delimiters=";,").delimiter
                except csv.Error:
                    delim = ";"
            return list(csv.DictReader(f, delimiter=delim))
    raise ValueError(f"Format non supporté : {path.suffix} (attendu .xlsx/.csv/.tsv)")


def load_products(path: str | Path) -> LoadResult:
    """Lit le fichier, mappe les colonnes et renvoie des `Product` normalisés."""
    path = Path(path)
    raw_rows = _read_rows(path)

    # Construire le mapping en-tête -> champ à partir de la 1re ligne de données.
    headers = list(raw_rows[0].keys()) if raw_rows else []
    mapping: dict[str, str] = {}
    unmapped: list[str] = []
    for h in headers:
        field = COLUMN_ALIASES.get(_normalize_header(h))
        if field:
            mapping[h] = field
        elif h:
            unmapped.append(h)

    products: list[Product] = []
    for i, row in enumerate(raw_rows, start=1):
        values: dict[str, object] = {}
        for h, field in mapping.items():
            if field not in values or values[field] in (None, ""):
                values[field] = row.get(h)

        product = Product(
            row_index=i,
            name=str(values.get("name") or "").strip(),
            brand=str(values.get("brand") or "").strip(),
            category=str(values.get("category") or "").strip(),
            price_internal=parse_price(values.get("price_internal")),
            seller_internal=str(values.get("seller_internal") or "").strip(),
            condition=_parse_condition(values.get("condition")),
            ean=validate_ean(str(values.get("ean") or "")),
        )
        products.append(product)

    return LoadResult(products=products, mapping=mapping, unmapped_headers=unmapped)

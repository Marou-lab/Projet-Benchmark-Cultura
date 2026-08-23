"""Export Excel du catalogue segmenté au schéma de sortie cible.

En Phase 1, seules les colonnes connues du fichier sont remplies ; les colonnes
concurrentielles (prix concurrent, écarts, confiance…) restent vides jusqu'aux
phases de collecte. Le statut initial découle du segment.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from ..models import OUTPUT_COLUMNS, Product, Segment
from ..validation.diagnostic import Diagnostic

# Statut initial par segment (avant toute collecte).
_STATUT_PAR_SEGMENT = {
    Segment.A: "À vérifier",       # benchmarkable mais rien encore collecté
    Segment.B: "Hors périmètre",   # cas particulier : traitement dédié
    Segment.C: "Exclu",            # non identifiable
}


def _row_for_product(p: Product) -> dict[str, object]:
    ean = p.ean
    statut = _STATUT_PAR_SEGMENT.get(p.segment, "À vérifier") if p.segment else "À vérifier"
    return {
        "EAN": ean.normalized if ean and ean.valid else (ean.raw if ean else ""),
        "Produit": p.name,
        "Marque": p.brand,
        "Catégorie": p.category,
        "Prix Cultura (réf. interne)": p.price_internal,
        "Vendeur Cultura": p.seller_internal,
        "Devise": "EUR",
        "Statut": statut,
        "Analyse métier": " | ".join(p.segment_reasons),
        "Segment": p.segment.value if p.segment else "",
    }


def write_workbook(products: list[Product], diag: Diagnostic, path: str | Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Feuille 1 : catalogue segmenté au schéma de sortie.
    ws = wb.active
    ws.title = "Catalogue segmenté"
    ws.append(OUTPUT_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for p in products:
        data = _row_for_product(p)
        ws.append([data.get(col, "") for col in OUTPUT_COLUMNS])
    ws.freeze_panes = "A2"

    # Feuille 2 : diagnostic (synthèse).
    ds = wb.create_sheet("Diagnostic")
    ds.append(["Indicateur", "Valeur"])
    for cell in ds[1]:
        cell.font = Font(bold=True)
    stats = [
        ("Produits (lignes)", diag.total),
        ("EAN valides", diag.ean_valid),
        ("EAN invalides/manquants", diag.ean_invalid_or_missing),
        ("EAN valides uniques", diag.unique_eans),
        ("Doublons EAN (nb codes)", len(diag.duplicate_eans)),
        ("Segment A (benchmarkable)", diag.segments.get("A", 0)),
        ("Segment B (cas particulier)", diag.segments.get("B", 0)),
        ("Segment C (non identifiable)", diag.segments.get("C", 0)),
        ("Prix réf. interne renseigné", diag.with_internal_price),
        ("Marques distinctes", diag.distinct_brands),
        ("Catégories distinctes", diag.distinct_categories),
        ("Vendeurs distincts", diag.distinct_sellers),
    ]
    for label, value in stats:
        ds.append([label, value])

    wb.save(path)
    return path

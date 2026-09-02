"""Rapport Excel de benchmark V1 — 3 onglets : Synthèse / Détail des offres / À vérifier."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

SYNTHESE_COLS = [
    "EAN", "Produit", "Métier", "VA TTC",
    "Statut Cdiscount", "Confiance",
    "Cultura 1P (€)", "Cultura meilleure globale livrée (€)",
    "Cdiscount meilleure (livrée) (€)", "Cdiscount vendeur", "Cdiscount 1P/3P",
    "Écart 1P↔concurrent €", "Écart 1P↔concurrent %",
    "Écart meilleures globales €", "Écart meilleures globales %",
    "Remarque / anomalie",
]

OFFRES_COLS = [
    "EAN", "Produit", "Plateforme", "Vendeur", "1P/3P",
    "Prix produit", "Livraison", "Prix total livré", "État",
    "Statut/Confiance", "URL source", "Date collecte",
]

ANOMALIES_COLS = ["EAN", "Produit", "Statut", "Raison", "Détail"]


def _sheet(wb: Workbook, title: str, cols: list[str], rows: list[dict], first: bool):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    ws.append(cols)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append([r.get(c, "") for c in cols])
    ws.freeze_panes = "A2"


def write_benchmark(path: str | Path, synthese: list[dict], offres: list[dict],
                    anomalies: list[dict]) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    _sheet(wb, "Synthèse", SYNTHESE_COLS, synthese, first=True)
    _sheet(wb, "Détail des offres", OFFRES_COLS, offres, first=False)
    _sheet(wb, "À vérifier - anomalies", ANOMALIES_COLS, anomalies, first=False)
    wb.save(path)
    return path
